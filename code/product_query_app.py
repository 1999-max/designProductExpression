import calendar
import json
import os
import queue
import re
import threading
import time
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lingxing_client import LingXingClient, RateLimitError


APP_TITLE = "产品表现 ASIN 查询"
APP_CONFIG_DIR = Path(os.getenv("APPDATA", Path.home())) / "ProductExpressionQuery"
APP_CONFIG_FILE = APP_CONFIG_DIR / "config.json"
PAGE_LENGTH = 1000
REQUEST_INTERVAL_SECONDS = 2
RATE_LIMIT_COOLDOWNS_SECONDS = [20, 45, 90]

MODE_BEFORE = "请求A日期7天前"
MODE_AFTER = "请求A日期7天后"
MODE_14_DAYS = "请求A日期14天内"
MODE_ALL = "请求全部"
MODE_OPTIONS = [MODE_BEFORE, MODE_AFTER, MODE_14_DAYS, MODE_ALL]

RANGE_LABELS = ["7天前", "7天后", "14天内"]
METRIC_ROWS = [
    ("ad_cvr", "广告CVR"),
    ("volume_cvr", "销量CVR"),
    ("cvr", "CVR"),
    ("ctr", "CTR"),
]

TABLE_COLUMNS = [
    ("metric", "指标", 160),
    ("before", "7天前", 150),
    ("after", "7天后", 150),
    ("within", "14天内", 150),
]

EXPORT_COLUMNS = [
    ("asin", "ASIN", 18),
    ("principal_names", "负责人", 18),
    ("metric", "指标", 18),
    ("before", "7天前", 16),
    ("after", "7天后", 16),
    ("within", "14天内", 16),
]


class QueryCancelled(Exception):
    pass


def parse_date(text):
    return datetime.strptime(text.strip(), "%Y-%m-%d").date()


def fmt_rate(value):
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.4f}"
    except (TypeError, ValueError):
        return str(value)


def parse_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def average(values):
    nums = [v for v in values if v is not None]
    if not nums:
        return None
    return sum(nums) / len(nums)


def date_ranges(date_a, mode):
    ranges = {
        MODE_BEFORE: [("7天前", date_a - timedelta(days=7), date_a)],
        MODE_AFTER: [("7天后", date_a, date_a + timedelta(days=7))],
        MODE_14_DAYS: [("14天内", date_a - timedelta(days=7), date_a + timedelta(days=7))],
        MODE_ALL: [
            ("7天前", date_a - timedelta(days=7), date_a),
            ("7天后", date_a, date_a + timedelta(days=7)),
            ("14天内", date_a - timedelta(days=7), date_a + timedelta(days=7)),
        ],
    }
    return ranges[mode]


def future_modes_locked(date_a):
    return date.today() < (date_a + timedelta(days=7))


def safe_filename_part(text):
    cleaned = re.sub(r"[\\/:*?\"<>|\s]+", "_", str(text).strip())
    return cleaned.strip("_") or "empty"


class CalendarPopup(tk.Toplevel):
    def __init__(self, master, initial_date, on_select):
        super().__init__(master)
        self.title("选择日期")
        self.resizable(False, False)
        self.configure(bg="#f5f7f3")
        self.transient(master)
        self.grab_set()
        self.on_select = on_select
        self.year = initial_date.year
        self.month = initial_date.month
        self._build()
        self._render()

    def _build(self):
        header = tk.Frame(self, bg="#f5f7f3")
        header.pack(fill="x", padx=12, pady=(12, 8))
        ttk.Button(header, text="<", width=3, command=self._prev_month).pack(side="left")
        self.title_label = tk.Label(header, bg="#f5f7f3", fg="#27322b", font=("Microsoft YaHei UI", 11, "bold"))
        self.title_label.pack(side="left", expand=True)
        ttk.Button(header, text=">", width=3, command=self._next_month).pack(side="right")
        self.days_frame = tk.Frame(self, bg="#f5f7f3")
        self.days_frame.pack(padx=12, pady=(0, 12))

    def _render(self):
        for child in self.days_frame.winfo_children():
            child.destroy()
        self.title_label.config(text=f"{self.year}年 {self.month:02d}月")
        for col, day_name in enumerate(["一", "二", "三", "四", "五", "六", "日"]):
            tk.Label(
                self.days_frame,
                text=day_name,
                width=4,
                bg="#f5f7f3",
                fg="#657069",
                font=("Microsoft YaHei UI", 9),
            ).grid(row=0, column=col, pady=(0, 4))
        for row, week in enumerate(calendar.monthcalendar(self.year, self.month), start=1):
            for col, day_num in enumerate(week):
                if day_num == 0:
                    tk.Label(self.days_frame, text="", width=4, bg="#f5f7f3").grid(row=row, column=col, pady=2)
                    continue
                selected = date(self.year, self.month, day_num)
                ttk.Button(
                    self.days_frame,
                    text=str(day_num),
                    width=4,
                    command=lambda d=selected: self._select(d),
                ).grid(row=row, column=col, padx=2, pady=2)

    def _select(self, selected):
        self.on_select(selected)
        self.destroy()

    def _prev_month(self):
        self.month -= 1
        if self.month == 0:
            self.month = 12
            self.year -= 1
        self._render()

    def _next_month(self):
        self.month += 1
        if self.month == 13:
            self.month = 1
            self.year += 1
        self._render()


class ProductQueryService:
    def __init__(self, log_callback, stop_event):
        self.client = LingXingClient()
        self.log = log_callback
        self.stop_event = stop_event
        self.last_request_time = 0

    def _check_cancelled(self):
        if self.stop_event.is_set():
            raise QueryCancelled("查询已中断")

    def _sleep(self, seconds):
        end_time = time.time() + seconds
        while time.time() < end_time:
            self._check_cancelled()
            time.sleep(min(0.25, end_time - time.time()))

    def _wait_interval(self):
        self._check_cancelled()
        elapsed = time.time() - self.last_request_time
        if elapsed < REQUEST_INTERVAL_SECONDS:
            self._sleep(REQUEST_INTERVAL_SECONDS - elapsed)
        self.last_request_time = time.time()

    def get_shops(self):
        self.log("正在获取店铺列表...")
        data = self.client.get("/erp/sc/data/seller/lists")
        shops = []
        for item in data or []:
            sid = item.get("sid")
            if sid:
                shops.append({"sid": str(sid), "name": item.get("name", "")})
        self.log(f"获取到 {len(shops)} 个店铺")
        return shops

    def fetch_shop_asin(self, shop, asin, start_date, end_date):
        records = []
        offset = 0
        while True:
            self._check_cancelled()
            self._wait_interval()
            body = {
                "offset": offset,
                "length": PAGE_LENGTH,
                "sort_field": "volume",
                "sort_type": "desc",
                "search_field": "asin",
                "search_value": [asin],
                "sid": str(shop["sid"]),
                "start_date": start_date.strftime("%Y-%m-%d"),
                "end_date": end_date.strftime("%Y-%m-%d"),
                "summary_field": "asin",
            }
            data = self.client.post("/bd/productPerformance/openApi/asinList", body=body)
            if not data:
                break
            page_records = data.get("list", [])
            records.extend(page_records)
            total = int(data.get("total") or 0)
            if offset + PAGE_LENGTH >= total or not page_records:
                break
            offset += PAGE_LENGTH
        return records

    def query(self, asin, selected_date, mode):
        ranges = date_ranges(selected_date, mode)
        shops = self.get_shops()
        range_data = self._empty_range_data(asin)
        failed_count = 0

        for range_label, start_date, end_date in ranges:
            self.log("")
            self.log(f"{range_label}: {start_date:%Y-%m-%d} ~ {end_date:%Y-%m-%d}")
            range_failed = []
            for index, shop in enumerate(shops, start=1):
                self._check_cancelled()
                shop_label = f"{shop['sid']} {shop.get('name', '')}".strip()
                self.log(f"请求店铺 {index}/{len(shops)}: {shop_label}")
                try:
                    records = self.fetch_shop_asin(shop, asin, start_date, end_date)
                    self._collect_records(range_data, range_label, records)
                    self.log(f"成功: {shop_label}，命中 {len(records)} 条")
                except RateLimitError as exc:
                    self.log(f"限流缓存: {shop_label}，{exc}")
                    range_failed.append((shop, exc))
                except Exception as exc:
                    failed_count += 1
                    self.log(f"失败: {shop_label}，{exc}")

            recovered, still_failed = self._retry_failed(range_data, range_label, start_date, end_date, asin, range_failed)
            failed_count += still_failed
            if recovered:
                self.log(f"{range_label}: 补跑成功 {recovered} 个店铺")

        rows = self._build_compare_rows(asin, range_data)
        return rows, failed_count

    def _retry_failed(self, range_data, range_label, start_date, end_date, asin, failed_items):
        recovered = 0
        pending = list(failed_items)
        for round_index, cooldown in enumerate(RATE_LIMIT_COOLDOWNS_SECONDS, start=1):
            if not pending:
                break
            self.log(f"补跑第 {round_index} 轮，冷却 {cooldown} 秒，待补跑 {len(pending)} 个店铺")
            self._sleep(cooldown)
            next_pending = []
            for shop, _exc in pending:
                self._check_cancelled()
                shop_label = f"{shop['sid']} {shop.get('name', '')}".strip()
                try:
                    records = self.fetch_shop_asin(shop, asin, start_date, end_date)
                    self._collect_records(range_data, range_label, records)
                    recovered += 1
                    self.log(f"补跑成功: {shop_label}，命中 {len(records)} 条")
                except RateLimitError as exc:
                    self.log(f"补跑仍限流: {shop_label}，{exc}")
                    next_pending.append((shop, exc))
                except Exception as exc:
                    self.log(f"补跑失败: {shop_label}，{exc}")
            pending = next_pending
        return recovered, len(pending)

    def _empty_range_data(self, asin):
        return {
            label: {
                "asin": asin,
                "principal_names": set(),
                "record_count": 0,
                "metrics": {key: [] for key, _label in METRIC_ROWS},
            }
            for label in RANGE_LABELS
        }

    def _collect_records(self, range_data, range_label, records):
        target = range_data[range_label]
        for item in records:
            target["record_count"] += 1
            principals = item.get("principal_names") or []
            if isinstance(principals, str):
                principals = [principals] if principals else []
            for principal in principals:
                if principal:
                    target["principal_names"].add(str(principal))
            for metric_key, _label in METRIC_ROWS:
                target["metrics"][metric_key].append(parse_float(item.get(metric_key)))

    def _build_compare_rows(self, asin, range_data):
        all_principals = set()
        for label in RANGE_LABELS:
            all_principals.update(range_data[label]["principal_names"])
        principal_text = ", ".join(sorted(all_principals))
        rows = []
        for metric_key, _metric_label in METRIC_ROWS:
            before = average(range_data["7天前"]["metrics"][metric_key])
            after = average(range_data["7天后"]["metrics"][metric_key])
            within = average(range_data["14天内"]["metrics"][metric_key])
            rows.append({
                "asin": asin,
                "principal_names": principal_text,
                "metric": _metric_label,
                "before": fmt_rate(before),
                "after": fmt_rate(after),
                "within": fmt_rate(within),
            })
        return rows


class ProductQueryApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1220x780")
        self.minsize(1060, 700)
        self.configure(bg="#eef2ed")
        self.result_rows = []
        self.queue = queue.Queue()
        self.worker = None
        self.stop_event = threading.Event()
        self.current_query = None
        self.cache_dir = self._ensure_cache_dir()
        self._style()
        self._build()
        self._update_mode_state()
        self.after(150, self._poll_queue)

    def _ensure_cache_dir(self):
        APP_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        if APP_CONFIG_FILE.exists():
            try:
                with open(APP_CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
                cache_dir = Path(config.get("cache_dir", ""))
                if cache_dir:
                    cache_dir.mkdir(parents=True, exist_ok=True)
                    return cache_dir
            except Exception:
                pass

        messagebox.showinfo("选择缓存位置", "请选择请求缓存保存位置。程序会在该位置创建或使用 json 文件夹。")
        selected = filedialog.askdirectory(title="选择请求缓存保存位置")
        if selected:
            base_dir = Path(selected)
            cache_dir = base_dir if base_dir.name.lower() == "json" else base_dir / "json"
        else:
            cache_dir = APP_CONFIG_DIR / "json"
            messagebox.showwarning("未选择缓存位置", f"未选择位置，已使用默认缓存文件夹：\n{cache_dir}")
        cache_dir.mkdir(parents=True, exist_ok=True)
        with open(APP_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({"cache_dir": str(cache_dir)}, f, ensure_ascii=False, indent=2)
        return cache_dir

    def _style(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", font=("Microsoft YaHei UI", 10))
        style.configure("TFrame", background="#eef2ed")
        style.configure("Surface.TFrame", background="#f8faf6")
        style.configure("Title.TLabel", background="#eef2ed", foreground="#1f2a24", font=("Microsoft YaHei UI", 19, "bold"))
        style.configure("Subtle.TLabel", background="#eef2ed", foreground="#627069", font=("Microsoft YaHei UI", 9))
        style.configure("Field.TLabel", background="#f8faf6", foreground="#44514a", font=("Microsoft YaHei UI", 9, "bold"))
        style.configure("Hint.TLabel", background="#f8faf6", foreground="#66746d", font=("Microsoft YaHei UI", 9))
        style.configure("Status.TLabel", background="#eef2ed", foreground="#53615a", font=("Microsoft YaHei UI", 9))
        style.configure("TButton", padding=(14, 8), background="#dde6de", foreground="#243029", borderwidth=0)
        style.map("TButton", background=[("active", "#d2dfd4"), ("disabled", "#e4e8e2")])
        style.configure("Accent.TButton", background="#2f6f55", foreground="#f7fbf6", borderwidth=0)
        style.map("Accent.TButton", background=[("active", "#285f49"), ("disabled", "#9eb3a9")])
        style.configure("Danger.TButton", background="#b95f4d", foreground="#fbf7f4", borderwidth=0)
        style.map("Danger.TButton", background=[("active", "#a64f40"), ("disabled", "#ddc1b9")])
        style.configure("TEntry", fieldbackground="#fbfcfa", bordercolor="#cbd6ce", lightcolor="#cbd6ce", darkcolor="#cbd6ce")
        style.configure("TCombobox", fieldbackground="#fbfcfa", bordercolor="#cbd6ce", arrowcolor="#3f4c45")
        style.configure("Treeview", background="#fbfcfa", fieldbackground="#fbfcfa", foreground="#26312b", rowheight=34, bordercolor="#d7ded8")
        style.configure("Treeview.Heading", background="#dfe8df", foreground="#2b362f", font=("Microsoft YaHei UI", 9, "bold"))

    def _build(self):
        outer = ttk.Frame(self, padding=24)
        outer.pack(fill="both", expand=True)

        header = ttk.Frame(outer)
        header.pack(fill="x", pady=(0, 18))
        ttk.Label(header, text="产品表现 ASIN 查询", style="Title.TLabel").pack(anchor="w")
        ttk.Label(header, text="单店铺请求，限流自动补跑，结果以指标对比表展示", style="Subtle.TLabel").pack(anchor="w", pady=(4, 0))

        body = ttk.Frame(outer)
        body.pack(fill="both", expand=True)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        controls = ttk.Frame(body, style="Surface.TFrame", padding=18)
        controls.grid(row=0, column=0, sticky="ns", padx=(0, 16))

        self.date_var = tk.StringVar(value=(date.today() - timedelta(days=7)).strftime("%Y-%m-%d"))
        self.mode_var = tk.StringVar(value=MODE_ALL)
        self.asin_var = tk.StringVar()

        self._field_label(controls, "日期A").pack(anchor="w")
        date_box = ttk.Frame(controls, style="Surface.TFrame")
        date_box.pack(fill="x", pady=(6, 18))
        self.date_entry = ttk.Entry(date_box, textvariable=self.date_var, width=18)
        self.date_entry.pack(side="left", fill="x", expand=True)
        ttk.Button(date_box, text="选择", command=self._open_calendar).pack(side="left", padx=(8, 0))
        self.date_entry.bind("<FocusOut>", lambda _event: self._update_mode_state())
        self.date_entry.bind("<Return>", lambda _event: self._update_mode_state())

        self._field_label(controls, "模式").pack(anchor="w")
        self.mode_combo = ttk.Combobox(controls, textvariable=self.mode_var, values=MODE_OPTIONS, width=26, state="readonly")
        self.mode_combo.pack(fill="x", pady=(6, 18))

        self._field_label(controls, "ASIN").pack(anchor="w")
        self.asin_entry = ttk.Entry(controls, textvariable=self.asin_var, width=28)
        self.asin_entry.pack(fill="x", pady=(6, 12))

        self.lock_label = ttk.Label(controls, text="", style="Hint.TLabel", wraplength=260)
        self.lock_label.pack(fill="x", pady=(0, 18))

        self.query_btn = ttk.Button(controls, text="查询", style="Accent.TButton", command=self._start_query)
        self.query_btn.pack(fill="x", pady=(0, 8))
        self.export_btn = ttk.Button(controls, text="导出", command=self._export, state="disabled")
        self.export_btn.pack(fill="x", pady=(0, 8))
        self.stop_btn = ttk.Button(controls, text="中断", style="Danger.TButton", command=self._stop_query, state="disabled")
        self.stop_btn.pack(fill="x")

        ttk.Label(
            controls,
            text="表格按 ASIN 聚合：4 个指标分别对比 7天前、7天后、14天内；多店铺命中时取有效值平均。",
            style="Hint.TLabel",
            wraplength=260,
        ).pack(fill="x", pady=(20, 0))

        right = ttk.Frame(body)
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(0, weight=1)
        right.rowconfigure(1, weight=0)
        right.columnconfigure(0, weight=1)

        result_frame = ttk.Frame(right, style="Surface.TFrame", padding=14)
        result_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 14))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(1, weight=1)

        summary = ttk.Frame(result_frame, style="Surface.TFrame")
        summary.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        summary.columnconfigure(1, weight=1)
        ttk.Label(summary, text="ASIN", style="Field.TLabel").grid(row=0, column=0, sticky="w")
        self.summary_asin_var = tk.StringVar(value="-")
        ttk.Label(summary, textvariable=self.summary_asin_var, style="Hint.TLabel").grid(row=0, column=1, sticky="w", padx=(10, 22))
        ttk.Label(summary, text="负责人", style="Field.TLabel").grid(row=0, column=2, sticky="w")
        self.summary_principal_var = tk.StringVar(value="-")
        ttk.Label(summary, textvariable=self.summary_principal_var, style="Hint.TLabel", wraplength=260).grid(row=0, column=3, sticky="w", padx=(10, 0))

        self.tree = ttk.Treeview(result_frame, columns=[c[0] for c in TABLE_COLUMNS], show="headings", height=8)
        for key, label, width in TABLE_COLUMNS:
            self.tree.heading(key, text=label)
            self.tree.column(key, width=width, minwidth=86, anchor="center", stretch=True)
        self.tree.grid(row=1, column=0, sticky="nsew")

        log_frame = ttk.Frame(right, style="Surface.TFrame", padding=12)
        log_frame.grid(row=1, column=0, sticky="ew")
        log_frame.columnconfigure(0, weight=1)
        ttk.Label(log_frame, text="运行日志", style="Field.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
        self.log_text = tk.Text(
            log_frame,
            height=7,
            bg="#17211c",
            fg="#dbe6dd",
            insertbackground="#dbe6dd",
            relief="flat",
            wrap="word",
            font=("Consolas", 10),
        )
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.grid(row=1, column=0, sticky="ew")
        log_scroll.grid(row=1, column=1, sticky="ns")

        footer = ttk.Frame(outer)
        footer.pack(fill="x", pady=(12, 0))
        self.progress = ttk.Progressbar(footer, mode="indeterminate", length=190)
        self.progress.pack(side="left")
        self.status_var = tk.StringVar(value="准备就绪")
        ttk.Label(footer, textvariable=self.status_var, style="Status.TLabel").pack(side="left", padx=(12, 0), fill="x", expand=True)

    def _field_label(self, master, text):
        return ttk.Label(master, text=text, style="Field.TLabel")

    def _open_calendar(self):
        try:
            initial = parse_date(self.date_var.get())
        except ValueError:
            initial = date.today()
        CalendarPopup(self, initial, self._set_date)

    def _set_date(self, selected):
        self.date_var.set(selected.strftime("%Y-%m-%d"))
        self._update_mode_state()

    def _cache_path(self, asin, selected_date, mode):
        name = "_".join([
            safe_filename_part(asin.upper()),
            selected_date.strftime("%Y%m%d"),
            safe_filename_part(mode),
        ])
        return self.cache_dir / f"{name}.json"

    def _load_cache(self, asin, selected_date, mode):
        path = self._cache_path(asin, selected_date, mode)
        if not path.exists():
            return None
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if data.get("asin", "").upper() != asin.upper():
            return None
        if data.get("date_a") != selected_date.strftime("%Y-%m-%d"):
            return None
        if data.get("mode") != mode:
            return None
        return data

    def _save_cache(self, asin, selected_date, mode, rows, failed_count):
        path = self._cache_path(asin, selected_date, mode)
        payload = {
            "asin": asin,
            "date_a": selected_date.strftime("%Y-%m-%d"),
            "mode": mode,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "failed_count": failed_count,
            "rows": rows,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return path

    def _update_mode_state(self):
        try:
            date_a = parse_date(self.date_var.get())
        except ValueError:
            self.lock_label.config(text="日期格式应为 YYYY-MM-DD")
            return
        if future_modes_locked(date_a):
            self.mode_combo.configure(values=[MODE_BEFORE])
            if self.mode_var.get() != MODE_BEFORE:
                self.mode_var.set(MODE_BEFORE)
            self.lock_label.config(text="当前还没到 A+7 天，7天后、14天内、全部模式暂不可请求。")
        else:
            self.mode_combo.configure(values=MODE_OPTIONS)
            if self.mode_var.get() not in MODE_OPTIONS:
                self.mode_var.set(MODE_ALL)
            self.lock_label.config(text="")

    def _start_query(self):
        if self.worker and self.worker.is_alive():
            return
        try:
            selected_date = parse_date(self.date_var.get())
        except ValueError:
            messagebox.showwarning("日期错误", "请输入或选择正确日期，格式为 YYYY-MM-DD")
            return
        self._update_mode_state()
        mode = self.mode_var.get()
        asin = self.asin_var.get().strip()
        if not asin:
            messagebox.showwarning("缺少 ASIN", "请输入要查询的 ASIN")
            return

        self.stop_event.clear()
        self.result_rows = []
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.summary_asin_var.set("-")
        self.summary_principal_var.set("-")
        self.log_text.delete("1.0", "end")
        self.export_btn.configure(state="disabled")
        self.current_query = {
            "asin": asin,
            "selected_date": selected_date,
            "mode": mode,
        }

        cached = self._load_cache(asin, selected_date, mode)
        if cached:
            self._append_log(f"缓存命中：{self._cache_path(asin, selected_date, mode)}")
            self._finish_query(cached.get("rows", []), int(cached.get("failed_count") or 0), save_cache=False)
            return

        self.query_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.progress.start(12)
        self.status_var.set("开始查询...")
        self._append_log(f"未命中缓存，开始请求接口。缓存目录：{self.cache_dir}")

        self.worker = threading.Thread(target=self._query_worker, args=(asin, selected_date, mode), daemon=True)
        self.worker.start()

    def _stop_query(self):
        if self.worker and self.worker.is_alive():
            self.stop_event.set()
            self._append_log("收到中断请求，正在停止当前查询...")
            self.status_var.set("正在中断...")
            self.stop_btn.configure(state="disabled")

    def _query_worker(self, asin, selected_date, mode):
        try:
            service = ProductQueryService(lambda msg: self.queue.put(("log", msg)), self.stop_event)
            rows, failed_count = service.query(asin, selected_date, mode)
            self.queue.put(("done", {"rows": rows, "failed_count": failed_count}))
        except QueryCancelled:
            self.queue.put(("cancelled", "查询已中断"))
        except Exception as exc:
            self.queue.put(("error", str(exc)))

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.queue.get_nowait()
                if kind == "log":
                    self._append_log(payload)
                    if payload:
                        self.status_var.set(payload)
                elif kind == "done":
                    self._finish_query(payload["rows"], payload["failed_count"])
                elif kind == "cancelled":
                    self._cancel_query(payload)
                elif kind == "error":
                    self._fail_query(payload)
        except queue.Empty:
            pass
        self.after(150, self._poll_queue)

    def _append_log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{timestamp}] {message}\n" if message else "\n"
        self.log_text.insert("end", line)
        self.log_text.see("end")

    def _finish_query(self, rows, failed_count, save_cache=True):
        self.progress.stop()
        self.query_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self.result_rows = rows
        if rows:
            self.summary_asin_var.set(rows[0].get("asin", "-") or "-")
            self.summary_principal_var.set(rows[0].get("principal_names", "-") or "-")
        for row in rows:
            values = [row.get(key, "") for key, _label, _width in TABLE_COLUMNS]
            self.tree.insert("", "end", values=values)
        self.export_btn.configure(state="normal" if rows else "disabled")
        self.status_var.set(f"查询完成，对比表 {len(rows)} 行，失败店铺 {failed_count} 个")
        self._append_log(f"查询完成，对比表 {len(rows)} 行，失败店铺 {failed_count} 个")
        if save_cache and self.current_query and rows:
            path = self._save_cache(
                self.current_query["asin"],
                self.current_query["selected_date"],
                self.current_query["mode"],
                rows,
                failed_count,
            )
            self._append_log(f"结果已保存到缓存：{path}")

    def _cancel_query(self, message):
        self.progress.stop()
        self.query_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self.status_var.set(message)
        self._append_log(message)

    def _fail_query(self, message):
        self.progress.stop()
        self.query_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self.status_var.set("查询失败")
        self._append_log(f"查询失败: {message}")
        messagebox.showerror("查询失败", message)

    def _export(self):
        if not self.result_rows:
            messagebox.showinfo("没有数据", "当前没有可导出的查询结果")
            return
        default_name = f"ASIN查询_{self.asin_var.get().strip()}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = filedialog.asksaveasfilename(
            title="导出结果",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 文件", "*.xlsx"), ("JSON 文件", "*.json")],
        )
        if not path:
            return
        output_path = Path(path)
        if output_path.suffix.lower() == ".json":
            self._write_json(output_path)
        else:
            self._write_excel(output_path)
        messagebox.showinfo("导出完成", f"已导出到：\n{path}")

    def _write_json(self, path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.result_rows, f, ensure_ascii=False, indent=2)

    def _write_excel(self, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "ASIN对比表"
        headers = [label for _key, label, _width in EXPORT_COLUMNS]
        ws.append(headers)
        for row in self.result_rows:
            ws.append([row.get(key, "") for key, _label, _width in EXPORT_COLUMNS])
        header_fill = PatternFill("solid", fgColor="DFE8DF")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="2B362F")
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for col_index, (_key, _label, width) in enumerate(EXPORT_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = width
        ws.freeze_panes = "A2"
        wb.save(path)


if __name__ == "__main__":
    app = ProductQueryApp()
    app.mainloop()
